"""英文句子视觉化语法标注编码器。

实现目标：
1) 使用 NLTK 分词 + 词性标注。
2) 为每个单词的每个字母输出 6 位数字编码：
   - 前三位：词性/字体粗细与灰度/下着重号。
   - 后三位：从句与非谓语边界/句子成分/语法辅助(被动、情态、虚拟等)。
3) 返回可直接用于后续 Word 样式渲染的数据结构和字符串。

说明：
- 由于自然语言存在歧义，部分“子类判定”采用可解释的启发式规则。
- 代码中将重复逻辑抽成独立方法，便于扩展。
"""

from __future__ import annotations

from dataclasses import dataclass
from typing import Dict, Iterable, List, Sequence, Tuple
import re
from pathlib import Path
from xml.sax.saxutils import escape
from zipfile import ZipFile, ZIP_DEFLATED

try:
    import nltk
    from nltk import pos_tag, sent_tokenize, word_tokenize
except Exception:  # 允许仅解码场景在无nltk环境下运行
    nltk = None
    pos_tag = None
    sent_tokenize = None
    word_tokenize = None


# -----------------------------
# 数据结构
# -----------------------------


@dataclass
class LetterCode:
    char: str
    code: str  # 6位字符串


@dataclass
class TokenEncoding:
    token: str
    pos: str
    letters: List[LetterCode]

    @property
    def compact(self) -> str:
        """形如 d000100o000100 的紧凑输出。"""
        return "".join(f"{item.char}{item.code}" for item in self.letters)


class VisualGrammarEncoder:
    """将句子编码为每字母 6 位数字。"""

    def __init__(self) -> None:
        # 允许在无 nltk 环境下仅使用解码或docx工具方法
        if nltk is not None:
            self._ensure_nltk_resources()

        # 前三位中的第一位（词性大类）
        self.first_digit_map: Dict[str, str] = {
            "VERB": "0",
            "NOUN": "1",
            "ADJ": "2",
            "PRON": "3",
            "ADV": "4",
            "DET": "5",
            "CONJ": "6",
            "UH": "7",
            "PREP": "8",
            "NUM": "9",
            "OTHER": "0",  # 无法识别时兜底
        }

        self.negative_words = {
            "not", "never", "no", "none", "neither", "nor", "hardly", "scarcely", "seldom",
            "few", "little", "barely", "without",
        }
        self.modal_words = {
            "can", "could", "may", "might", "must", "shall", "should", "will", "would", "ought",
        }
        self.future_modals = {"will", "shall"}
        self.future_in_past_modals = {"would", "should"}
        self.copulas = {"am", "is", "are", "was", "were", "be", "been", "being"}
        self.aux_verbs = {
            "am", "is", "are", "was", "were", "be", "been", "being",
            "do", "does", "did", "have", "has", "had",
        }
        self.causative_verbs = {"let", "make", "have", "get", "help"}

        self.subordinators = {
            "that", "which", "who", "whom", "whose", "when", "where", "why", "how",
            "if", "whether", "because", "although", "though", "while", "since", "before", "after",
            "unless", "until", "once",
        }

        # 名词语块规则（[NN] 表示 NN/NNP/NNS/NNPS）
        self.noun_phrase_pattern_specs: List[str] = [
            "DT_JJ_NNP_NNP_POS_[NN]", "NNP_NNP_POS_[NN]", "PRP$_NN_POS_[NN]", "CD_NNS_POS_[NN]",
            "DT_NNP_NNP_[NN]", "DT_NNP_POS_[NN]", "DT_NNS_POS_[NN]", "JJ_NNP_POS_[NN]",
            "NN_NNS_POS_[NN]", "NNP_NNP_POS_JJ_[NN]", "DT_JJ_NNP_[NN]", "DT_NN_POS_[NN]",
            "DT_RBS_JJ_[NN]", "DT_RBS_NN_[NN]", "DT_RBS_[NN]_NN", "JJ_NNP_NN_[NN]",
            "NN_NN_POS_[NN]", "[NN]_NNS", "DT_JJ_JJ_[NN]", "DT_JJ_NN_[NN]", "DT_RB_JJ_[NN]",
            "RB_JJ_NN_[NN]", "NNP_POS_[NN]", "PRP$_DT_[NN]", "PRP$_JJ_[NN]", "PRP$_NN_[NN]",
            "DT_JJR_[NN]", "DT_JJS_[NN]", "DT_NNP_[NN]", "NN_POS_[NN]", "NNP_JJ_[NN]",
            "PDT_DT_[NN]", "CD_JJ_[NN]", "DT_DT_[NN]", "DT_JJ_[NN]", "DT_NN_[NN]", "NN_DT_[NN]",
            "RB_CD_[NN]", "RB_JJ_[NN]", "PRP$_[NN]", "JJR_[NN]", "NNP_[NN]", "NNS_[NN]",
            "POS_[NN]", "CD_[NN]", "DT_[NN]", "JJ_[NN]", "NN_[NN]", "[NN]", "DT_NNP_POS_[NN]_NN",
            "DT_JJ_CD_NNS", "DT_JJ_NNP_[NN]_NN", "DT_JJ_NNP_NN_[NN]", "DT_NN_POS_[NN]_NNS",
            "DT_JJ_NNP_NN_[NN]", "DT_JJ_NNP_[NN]_NN", "JJ_NNP_[NN]_NN", "DT_NN_NN_POS_[NN]",
            "DT_RB_JJ_[NN]_NN", "JJ_NNS", "CD_NNS", "DT_RB_JJ_NN_[NN]", "DT_RB_JJ_[NN]_NN",
            "DT_NNP_POS_[NN]", "NNP_NNP_POS_[NN]", "NNP_POS_[NN]_NNS", "DT_NNP_POS_[NN]_NN",
            "NNP_POS_[NN]_NN", "JJ_NNP_POS_[NN]", "DT_JJ_NNP_NNP_POS_[NN]", "PRP$_JJ_[NN]_NNP",
            "PRP$_JJ_[NN]_NN", "DT_JJ_[NN]_NN", "JJ_CC_JJ_NN", "JJ_,_JJ_CC_JJ_NN",
        ]
        self._noun_tag_set = {"NN", "NNP", "NNS", "NNPS"}
        self._compiled_noun_phrase_patterns = [
            (spec, self._compile_noun_pattern(spec)) for spec in self.noun_phrase_pattern_specs
        ]
        self._compiled_noun_phrase_patterns.sort(key=lambda item: len(item[1]), reverse=True)

    # -----------------------------
    # 公共方法
    # -----------------------------

    def encode_sentence(self, sentence: str) -> List[TokenEncoding]:
        if word_tokenize is None or pos_tag is None:
            raise RuntimeError("NLTK 不可用：请先安装 nltk 后再进行编码。")
        tokens = word_tokenize(sentence)
        tagged = self._tag_tokens(tokens)

        # 先计算“后3位”的句法辅助上下文
        clause_marks = self._detect_clause_nonfinite_marks(tagged)
        component_marks = self._detect_sentence_components(tagged)
        grammar_marks = self._detect_grammar_aux_marks(tagged)

        results: List[TokenEncoding] = []
        for i, (token, pos) in enumerate(tagged):
            letters = self._encode_token_letters(
                token=token,
                pos=pos,
                tagged=tagged,
                index=i,
                clause_mark=clause_marks[i],
                component_mark=component_marks[i],
                grammar_mark=grammar_marks[i],
            )
            results.append(TokenEncoding(token=token, pos=pos, letters=letters))
        return results

    def get_noun_phrases(self, text: str) -> Dict[str, List[Dict[str, object]]]:
        """Get noun phrases, chunk POS patterns and single-noun context combinations."""
        if sent_tokenize is None or word_tokenize is None or pos_tag is None:
            raise RuntimeError("NLTK is unavailable: install nltk before noun phrase extraction.")

        multiword_chunks: List[Dict[str, object]] = []
        single_nouns_with_context: List[Dict[str, object]] = []

        for sent_index, sentence in enumerate(sent_tokenize(text)):
            tagged = self._tag_tokens(word_tokenize(sentence))
            occupied = [False] * len(tagged)
            noun_indexes = [idx for idx, (_, pos) in enumerate(tagged) if pos in self._noun_tag_set]

            i = 0
            while i < len(tagged):
                match = self._match_noun_pattern_at(tagged, i)
                if not match:
                    i += 1
                    continue

                pattern, span = match
                end = i + span

                # 必须包含名词，且仅统计 2 词及以上名词块
                has_noun = any(j in noun_indexes for j in range(i, end))
                if span >= 2 and has_noun:
                    for j in range(i, end):
                        occupied[j] = True
                    tags = [pos for _, pos in tagged[i:end]]
                    multiword_chunks.append(
                        {
                            "sentence_index": sent_index,
                            "start": i,
                            "end": end - 1,
                            "text": " ".join(tok for tok, _ in tagged[i:end]),
                            "tokens": [tok for tok, _ in tagged[i:end]],
                            "tags": tags,
                            "pattern": pattern,
                            "pos_pattern": "_".join(tags),
                        }
                    )
                    i = end
                else:
                    i += 1

            for idx in noun_indexes:
                tok, pos = tagged[idx]
                if occupied[idx]:
                    continue

                # 第二类：未构成 2+ 词名词块的单个名词，记录其相邻前后词性。
                prev_pos = tagged[idx - 1][1] if idx - 1 >= 0 else "<BOS>"
                next_pos = tagged[idx + 1][1] if idx + 1 < len(tagged) else "<EOS>"

                single_nouns_with_context.append(
                    {
                        "sentence_index": sent_index,
                        "index": idx,
                        "token": tok,
                        "noun_tag": pos,
                        "context_pattern": f"{prev_pos}_{pos}_{next_pos}",
                        "prev_tag": prev_pos,
                        "next_tag": next_pos,
                    }
                )

        return {
            "multiword_chunks": multiword_chunks,
            "single_nouns_with_context": single_nouns_with_context,
            "multiword_pos_patterns": [chunk["pos_pattern"] for chunk in multiword_chunks],
            "single_noun_context_patterns": [item["context_pattern"] for item in single_nouns_with_context],
        }

    def extract_noun_phrase_chunks(self, text: str) -> Dict[str, List[Dict[str, object]]]:
        """Backward-compatible alias for get_noun_phrases."""
        return self.get_noun_phrases(text)

    def get_labeled_noun_results(self, text: str) -> Dict[str, List[Dict[str, str]]]:
        """返回适合展示/导出的两类名词分析结果。"""
        raw_result = self.get_noun_phrases(text)

        labeled_multiword: List[Dict[str, str]] = []
        for chunk in raw_result["multiword_chunks"]:
            labeled_multiword.append(
                {
                    "标注类型": "2词及以上名词块模式",
                    "句子序号": f"S{chunk['sentence_index'] + 1}",
                    "名词块文本": str(chunk["text"]),
                    "词性组合模式": str(chunk["pos_pattern"]),
                }
            )

        labeled_single: List[Dict[str, str]] = []
        for single in raw_result["single_nouns_with_context"]:
            labeled_single.append(
                {
                    "标注类型": "单个名词+前后词性搭配组合",
                    "句子序号": f"S{single['sentence_index'] + 1}",
                    "单个名词": str(single["token"]),
                    "前后词性搭配模式": str(single["context_pattern"]),
                }
            )

        return {
            "labeled_multiword": labeled_multiword,
            "labeled_single": labeled_single,
        }

    def export_noun_results_to_excel(
        self,
        text: str,
        output_excel: str = "名词块分析结果.xlsx",
        multi_label: str = "2词及以上名词块模式",
        single_label: str = "单个名词+前后词性搭配组合",
    ) -> str:
        """导出名词块分析到 Excel（两个工作表）。"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Font, PatternFill
        except ImportError as exc:
            raise RuntimeError("请先安装 openpyxl：pip install openpyxl") from exc

        labeled = self.get_labeled_noun_results(text)

        wb = Workbook()

        ws_multi = wb.active
        ws_multi.title = multi_label[:31] or "Multiword"
        headers_multi = ["句子序号", "名词块文本", "词性组合模式"]
        for col, header in enumerate(headers_multi, 1):
            cell = ws_multi.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        for row, item in enumerate(labeled["labeled_multiword"], 2):
            ws_multi.cell(row=row, column=1, value=item["句子序号"])
            ws_multi.cell(row=row, column=2, value=item["名词块文本"])
            ws_multi.cell(row=row, column=3, value=item["词性组合模式"])

        if not labeled["labeled_multiword"]:
            ws_multi.cell(row=2, column=1, value="(none)")

        ws_multi.column_dimensions["A"].width = 12
        ws_multi.column_dimensions["B"].width = 36
        ws_multi.column_dimensions["C"].width = 26

        ws_single = wb.create_sheet(title=single_label[:31] or "Single")
        headers_single = ["句子序号", "单个名词", "前后词性搭配模式"]
        for col, header in enumerate(headers_single, 1):
            cell = ws_single.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        for row, item in enumerate(labeled["labeled_single"], 2):
            ws_single.cell(row=row, column=1, value=item["句子序号"])
            ws_single.cell(row=row, column=2, value=item["单个名词"])
            ws_single.cell(row=row, column=3, value=item["前后词性搭配模式"])

        if not labeled["labeled_single"]:
            ws_single.cell(row=2, column=1, value="(none)")

        ws_single.column_dimensions["A"].width = 12
        ws_single.column_dimensions["B"].width = 18
        ws_single.column_dimensions["C"].width = 32

        output_path = Path(output_excel)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_path)
        return str(output_path.resolve())


    def format_noun_phrase_report(self, text: str) -> List[str]:
        """格式化名词识别结果，便于直接打印查看。"""
        result = self.get_noun_phrases(text)
        lines: List[str] = ["=== Noun Phrase Report ===", "[2+词名词块模式]"]

        if result["multiword_chunks"]:
            for chunk in result["multiword_chunks"]:
                lines.append(f"S{chunk['sentence_index']} {chunk['text']} -> {chunk['pos_pattern']}")
        else:
            lines.append("(none)")

        lines.append("[单个名词前后词性搭配]")
        if result["single_nouns_with_context"]:
            for item in result["single_nouns_with_context"]:
                lines.append(
                    f"S{item['sentence_index']} {item['token']} -> {item['context_pattern']}"
                )
        else:
            lines.append("(none)")

        return lines

    def encode_sentence_as_dict(self, sentence: str) -> List[Dict[str, object]]:
        """便于 JSON 序列化。"""
        encoded = self.encode_sentence(sentence)
        return [
            {
                "token": item.token,
                "pos": item.pos,
                "compact": item.compact,
                "letters": [{"char": l.char, "code": l.code} for l in item.letters],
            }
            for item in encoded
        ]

    def encode_text(self, text: str) -> List[TokenEncoding]:
        """对整段文本进行编码（按句切分，逐句编码并合并）。"""
        if sent_tokenize is None:
            raise RuntimeError("NLTK 不可用：请先安装 nltk 后再进行编码。")
        all_tokens: List[TokenEncoding] = []
        for para in text.splitlines():
            para = para.strip()
            if not para:
                continue
            for sentence in sent_tokenize(para):
                all_tokens.extend(self.encode_sentence(sentence))
        return all_tokens

    def decode_compact_token(self, compact: str) -> str:
        """把形如 d000100o000100 的紧凑编码还原为原单词。"""
        i = 0
        chars: List[str] = []
        while i < len(compact):
            ch = compact[i]
            # 数字 token 在编码时可能原样返回
            if i + 7 > len(compact) or not compact[i + 1:i + 7].isdigit():
                chars.append(ch)
                i += 1
            else:
                chars.append(ch)
                i += 7
        return "".join(chars)

    def decode_compact_text(self, text: str) -> str:
        """将多行 token/POS: compact 输出反解回可读文本。"""
        lines: List[str] = []
        for raw in text.splitlines():
            line = raw.strip()
            if not line:
                lines.append("")
                continue
            if ":" not in line:
                lines.append(line)
                continue
            _, right = line.split(":", 1)
            decoded_parts: List[str] = []
            for piece in right.strip().split():
                decoded_parts.append(self.decode_compact_token(piece))
            lines.append(" ".join(decoded_parts))
        return "\n".join(lines)


    def encode_text_lines(self, text: str) -> List[str]:
        """返回适合界面显示的逐 token 编码行。"""
        lines: List[str] = []
        for item in self.encode_text(text):
            if item.letters and item.letters[0].code == "":
                lines.append(f"{item.token}/{item.pos}: {item.token}")
            else:
                lines.append(f"{item.token}/{item.pos}: {item.compact}")
        return lines

    def save_encoded_text_to_word(self, text: str, output_docx: str = "encoded_text_output.docx") -> str:
        """将整段文本编码结果保存到 Word。"""
        encoded = self.encode_text(text)
        lines = ["Encoded text tokens:"]
        for item in encoded:
            if item.letters and item.letters[0].code == "":
                lines.append(f"{item.token}/{item.pos}: {item.token}")
            else:
                lines.append(f"{item.token}/{item.pos}: {item.compact}")
        return self._write_simple_docx(lines, output_docx)

    def save_sentence_to_word(self, sentence: str, output_docx: str = "encoded_output.docx") -> str:
        """将句子的编码结果保存到 Word 文档(.docx)。

        文档内容格式：
        - 第一行：Original sentence
        - 后续每行：token/POS: c123456h123456...
        """
        encoded = self.encode_sentence(sentence)
        lines = [f"Original sentence: {sentence}", "", "Encoded tokens:"]
        for item in encoded:
            if item.letters and item.letters[0].code == "":
                lines.append(f"{item.token}/{item.pos}: {item.token}")
            else:
                lines.append(f"{item.token}/{item.pos}: {item.compact}")
        return self._write_simple_docx(lines, output_docx)

    # -----------------------------
    # NLTK 准备
    # -----------------------------

    def _ensure_nltk_resources(self) -> None:
        if nltk is None:
            raise RuntimeError("NLTK 未安装，编码功能不可用；但解码和docx写入功能仍可单独使用。")
        resources = [
            ("tokenizers/punkt", "punkt"),
            ("tokenizers/punkt_tab", "punkt_tab"),
            ("taggers/averaged_perceptron_tagger", "averaged_perceptron_tagger"),
            ("taggers/averaged_perceptron_tagger_eng", "averaged_perceptron_tagger_eng"),
        ]
        for resource_path, package in resources:
            try:
                nltk.data.find(resource_path)
            except LookupError:
                nltk.download(package, quiet=True)

    def _tag_tokens(self, tokens: Sequence[str]) -> List[Tuple[str, str]]:
        """先用 NLTK 标注，再做可解释的规则纠偏，提高整段文本稳定性。"""
        tagged = list(pos_tag(tokens))
        return self._retag_with_rules(tagged)

    def _retag_with_rules(self, tagged: List[Tuple[str, str]]) -> List[Tuple[str, str]]:
        """对 NLTK 结果做轻量后处理，缓解常见误标。"""
        force_tags = {
            "a": "DT", "an": "DT", "the": "DT",
            "this": "DT", "that": "DT", "these": "DT", "those": "DT",
            "my": "PRP$", "your": "PRP$", "our": "PRP$", "their": "PRP$", "its": "PRP$",
            "with": "IN", "for": "IN", "from": "IN", "of": "IN", "at": "IN", "by": "IN", "about": "IN",
            "and": "CC", "or": "CC", "but": "CC", "nor": "CC",
            "can": "MD", "could": "MD", "may": "MD", "might": "MD", "must": "MD",
            "shall": "MD", "should": "MD", "will": "MD", "would": "MD",
            "not": "RB", "never": "RB",
        }

        fixed = list(tagged)
        for i, (tok, pos) in enumerate(fixed):
            low = tok.lower()
            if low in force_tags:
                fixed[i] = (tok, force_tags[low])
                continue

            # to + VB 统一看作不定式标记
            if low == "to" and i + 1 < len(fixed) and fixed[i + 1][1].startswith("VB"):
                fixed[i] = (tok, "TO")
                continue

            # 纯数字统一 CD
            if tok.isdigit():
                fixed[i] = (tok, "CD")

            # 句首称呼语，如 "Tom," 倾向 NNP
            if i == 0 and i + 1 < len(fixed) and fixed[i + 1][0] == "," and pos in {"NN", "NNS"}:
                fixed[i] = (tok, "NNP")

        return fixed

    def _compile_noun_pattern(self, pattern: str) -> List[set[str]]:
        compiled: List[set[str]] = []
        for part in pattern.split("_"):
            if part == "[NN]":
                compiled.append(set(self._noun_tag_set))
            else:
                compiled.append({part})
        return compiled

    def _match_noun_pattern_at(
        self,
        tagged: Sequence[Tuple[str, str]],
        start: int,
    ) -> Tuple[str, int] | None:
        for spec, compiled in self._compiled_noun_phrase_patterns:
            span = len(compiled)
            if start + span > len(tagged):
                continue
            ok = True
            for i, options in enumerate(compiled):
                if tagged[start + i][1] not in options:
                    ok = False
                    break
            if ok:
                return spec, span
        return None

    def _write_simple_docx(self, lines: Sequence[str], output_docx: str) -> str:
        """零第三方依赖写入最小可打开的 .docx 文件。"""
        output = Path(output_docx)
        if output.suffix.lower() != ".docx":
            output = output.with_suffix(".docx")

        content_types = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">
  <Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>
  <Default Extension="xml" ContentType="application/xml"/>
  <Override PartName="/word/document.xml" ContentType="application/vnd.openxmlformats-officedocument.wordprocessingml.document.main+xml"/>
</Types>"""

        rels = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
  <Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="word/document.xml"/>
</Relationships>"""

        body = []
        for line in lines:
            text = escape(line)
            paragraph = (
                "<w:p><w:r><w:t xml:space=\"preserve\">"
                f"{text}"
                "</w:t></w:r></w:p>"
            )
            body.append(paragraph)

        document = f"""<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<w:document xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main">
  <w:body>
    {''.join(body)}
    <w:sectPr/>
  </w:body>
</w:document>"""

        output.parent.mkdir(parents=True, exist_ok=True)
        with ZipFile(output, "w", compression=ZIP_DEFLATED) as zf:
            zf.writestr("[Content_Types].xml", content_types)
            zf.writestr("_rels/.rels", rels)
            zf.writestr("word/document.xml", document)

        return str(output.resolve())

    # -----------------------------
    # 编码核心
    # -----------------------------

    def _encode_token_letters(
        self,
        token: str,
        pos: str,
        tagged: Sequence[Tuple[str, str]],
        index: int,
        clause_mark: str,
        component_mark: str,
        grammar_mark: str,
    ) -> List[LetterCode]:
        # 纯数字 CD：按需求原样返回
        if pos == "CD" and token.isdigit():
            return [LetterCode(ch, "") for ch in token]

        chars = list(token)
        n = len(chars)
        digits = [["0", "0", "0"] for _ in range(n)]

        # 默认第一位
        first = self._category_first_digit(pos)
        for d in digits:
            d[0] = first

        # 基础前三位规则
        self._apply_pos_rules(token, pos, tagged, index, digits)

        # 拼接后3位
        final: List[LetterCode] = []
        for ch, d in zip(chars, digits):
            six = "".join(d) + clause_mark + component_mark + grammar_mark
            final.append(LetterCode(ch, six))
        return final

    def _category_first_digit(self, pos: str) -> str:
        if pos.startswith("VB") or pos == "MD":
            return self.first_digit_map["VERB"]
        if pos in {"NN", "NNS", "NNP", "NNPS", "POS"}:
            return self.first_digit_map["NOUN"]
        if pos.startswith("JJ"):
            return self.first_digit_map["ADJ"]
        if pos in {"PRP", "PRP$", "WP", "WP$", "WDT"}:
            return self.first_digit_map["PRON"]
        if pos in {"RB", "RBR", "RBS", "RP", "WRB", "EX"}:
            return self.first_digit_map["ADV"]
        if pos in {"DT", "PDT"}:
            return self.first_digit_map["DET"]
        if pos in {"CC"}:
            return self.first_digit_map["CONJ"]
        if pos == "UH":
            return self.first_digit_map["UH"]
        if pos == "TO":
            return self.first_digit_map["PREP"]
        if pos == "IN":
            return self.first_digit_map["PREP"]
        if pos == "CD":
            return self.first_digit_map["NUM"]
        return self.first_digit_map["OTHER"]

    # -----------------------------
    # 规则方法（可复用）
    # -----------------------------

    def _apply_pos_rules(
        self,
        token: str,
        pos: str,
        tagged: Sequence[Tuple[str, str]],
        index: int,
        digits: List[List[str]],
    ) -> None:
        lower = token.lower()

        if pos.startswith("VB") or pos == "MD" or (pos == "TO"):
            self._apply_verb_rules(lower, pos, tagged, index, digits)
            return

        if pos in {"NN", "NNS", "NNP", "NNPS", "POS"}:
            self._apply_noun_rules(lower, pos, digits)
            return

        if pos.startswith("JJ"):
            self._apply_adj_rules(lower, pos, digits)
            return

        if pos in {"PRP", "PRP$", "WP", "WP$", "WDT"}:
            self._apply_pron_rules(lower, pos, digits)
            return

        if pos in {"RB", "RBR", "RBS", "RP", "WRB", "EX"}:
            self._apply_adv_rules(lower, pos, digits)
            return

        if pos in {"DT", "PDT"}:
            self._apply_det_rules(lower, digits)
            return

        if pos in {"CC", "IN"}:
            self._apply_conj_or_prep_rules(lower, pos, digits)
            return

        if pos == "UH":
            self._apply_uh_rules(lower, digits)
            return

        if pos == "CD":
            self._apply_num_rules(lower, digits)
            return

        if pos in {"LS", "SYM", "FW"}:
            for d in digits:
                d[1], d[2] = "0", "0"

    def _apply_emphasis(self, digits: List[List[str]], indices: Iterable[int], second: str | None = None, third: str | None = None) -> None:
        for i in indices:
            if 0 <= i < len(digits):
                if second is not None:
                    digits[i][1] = second
                if third is not None:
                    digits[i][2] = third

    def _middle_index(self, n: int) -> int:
        return n // 2 if n > 0 else 0

    def _apply_verb_rules(self, lower: str, pos: str, tagged: Sequence[Tuple[str, str]], index: int, digits: List[List[str]]) -> None:
        n = len(digits)

        # TO: 介词to / 不定式to
        if pos == "TO":
            next_pos = tagged[index + 1][1] if index + 1 < len(tagged) else ""
            if next_pos == "VB":
                for d in digits:
                    d[0] = "0"  # 视为动词原形引导
            else:
                for d in digits:
                    d[0] = "8"  # 介词
            return

        # VB 原形
        if pos == "VB":
            if lower in self.causative_verbs and n >= 2:
                self._apply_emphasis(digits, [self._middle_index(n)], second="1")
            return

        # VBD 过去式：后两个下着重号=1
        if pos == "VBD":
            self._apply_emphasis(digits, range(max(0, n - 2), n), third="1")
            if lower in self.aux_verbs and n >= 1:
                self._apply_emphasis(digits, [0], second="1")
            return

        # VBG 现在分词：后三个加粗；若动名词 -> 后四个加粗
        if pos == "VBG":
            is_gerund = self._is_gerund_context(tagged, index)
            k = 4 if is_gerund else 3
            self._apply_emphasis(digits, range(max(0, n - k), n), second="1")
            return

        # VBN 过去分词：后两个加粗
        if pos == "VBN":
            self._apply_emphasis(digits, range(max(0, n - 2), n), second="1")
            return

        # VBP 非三单现在：默认全部加粗
        if pos == "VBP":
            self._apply_emphasis(digits, range(n), second="1")
            if lower in self.aux_verbs:
                if n <= 3:
                    # 短助动词只突出助动属性：首字母加粗，其余恢复常规
                    for i in range(1, n):
                        digits[i][1] = "0"
                    self._apply_emphasis(digits, [0], second="1")
                else:
                    self._apply_emphasis(digits, [0], second="1")
            mid = self._middle_index(n) if n > 2 else 0
            self._apply_emphasis(digits, [mid], second="1")
            return

        # VBZ 三单现在：末字母加粗 + 中间突出
        if pos == "VBZ":
            if n:
                self._apply_emphasis(digits, [n - 1], second="1")
            if lower in self.aux_verbs and n > 2:
                self._apply_emphasis(digits, [0], second="1")
            mid = self._middle_index(n) if n > 2 else 0
            self._apply_emphasis(digits, [mid], second="1")
            return

        # MD 情态动词
        if pos == "MD":
            # 前两位加粗
            self._apply_emphasis(digits, [0, 1], second="1")
            if lower in self.future_modals:
                # 前三位下着重号：将来
                self._apply_emphasis(digits, [0, 1, 2], third="1")
            elif lower in self.future_in_past_modals:
                # 前两后两下着重号：过去将来
                self._apply_emphasis(digits, [0, 1, n - 2, n - 1], third="1")
            else:
                # 一般情态：中间位下着重号
                self._apply_emphasis(digits, [self._middle_index(n)], third="1")
            return

    def _apply_noun_rules(self, lower: str, pos: str, digits: List[List[str]]) -> None:
        n = len(digits)
        if pos in {"NNS", "NNPS"} and n:
            self._apply_emphasis(digits, [n - 1], second="1")
        elif pos == "NN":
            # 启发式：不可数常见词
            if lower in {"water", "milk", "money", "information", "advice", "furniture", "news"}:
                self._apply_emphasis(digits, [n - 2, n - 1], second="1")
            # 抽象名词后缀
            elif re.search(r"(tion|sion|ness|ity|ment|ship|ism|age)$", lower):
                self._apply_emphasis(digits, [self._middle_index(n)], second="1")
            # 集体名词示例
            elif lower in {"team", "family", "group", "government", "staff", "audience"}:
                self._apply_emphasis(digits, [0, 1, 2], second="1")
        elif pos == "POS":
            for d in digits:
                d[1] = "2"

    def _apply_adj_rules(self, lower: str, pos: str, digits: List[List[str]]) -> None:
        n = len(digits)
        if lower in self.negative_words:
            self._apply_emphasis(digits, range(n), second="2")
            if n >= 2:
                self._apply_emphasis(digits, [self._middle_index(n)], second="1")
            return

        if pos == "JJR":
            self._apply_emphasis(digits, [n - 2, n - 1], second="1")
        elif pos == "JJS":
            self._apply_emphasis(digits, [n - 3, n - 2, n - 1], second="1")

    def _apply_pron_rules(self, lower: str, pos: str, digits: List[List[str]]) -> None:
        n = len(digits)
        object_pronouns = {"me", "him", "her", "us", "them", "you", "whom"}
        possessive_independent = {"mine", "yours", "his", "hers", "ours", "theirs"}

        if pos == "PRP":
            if lower in object_pronouns and n:
                self._apply_emphasis(digits, [n - 1], second="2")
        elif pos == "PRP$":
            if n:
                self._apply_emphasis(digits, [0], second="2")
            if lower in possessive_independent and n:
                self._apply_emphasis(digits, [self._middle_index(n)], second="2")
        elif pos == "WDT":
            self._apply_emphasis(digits, [0, 1], second="2")
        elif pos in {"WP", "WP$"}:
            if n >= 2:
                self._apply_emphasis(digits, [n - 2, n - 1], second="2")

    def _apply_adv_rules(self, lower: str, pos: str, digits: List[List[str]]) -> None:
        n = len(digits)
        if pos == "RB":
            if lower in self.negative_words:
                self._apply_emphasis(digits, range(n), second="2")
            elif lower in {"as", "so", "equally"} and n:
                self._apply_emphasis(digits, [n - 1], second="2")
        elif pos == "RBR":
            self._apply_emphasis(digits, [n - 2, n - 1], second="2")
        elif pos == "RBS":
            self._apply_emphasis(digits, [n - 3, n - 2, n - 1], second="2")
        elif pos == "RP" and n:
            self._apply_emphasis(digits, [0], second="2")
        elif pos == "WRB":
            self._apply_emphasis(digits, [0, 1], second="2")

    def _apply_det_rules(self, lower: str, digits: List[List[str]]) -> None:
        n = len(digits)
        if lower in self.negative_words or lower in {"no", "neither", "none", "few", "little"}:
            # 按需求：否定意义限定词以 410 风格呈现 -> 第一位属于副词类4
            for d in digits:
                d[0], d[1], d[2] = "4", "1", "0"
        else:
            for d in digits:
                d[0], d[1], d[2] = "5", "0", "0"

    def _apply_conj_or_prep_rules(self, lower: str, pos: str, digits: List[List[str]]) -> None:
        n = len(digits)
        paired = {"either", "neither", "nor", "both", "and", "not", "only", "or"}

        if pos == "CC":
            for d in digits:
                d[0], d[1], d[2] = "6", "0", "0"
            if lower in paired and n:
                self._apply_emphasis(digits, [0], second="2")
            return

        # IN: 可能是介词或从属连词
        if lower in self.subordinators:
            for d in digits:
                d[0], d[1], d[2] = "6", "0", "0"
            if n:
                self._apply_emphasis(digits, [n - 1], second="2")
        else:
            for d in digits:
                d[0], d[1], d[2] = "8", "0", "0"

    def _apply_uh_rules(self, lower: str, digits: List[List[str]]) -> None:
        greetings = {"morning", "evening", "hello", "hi"}
        if lower in greetings:
            # 问候语：末字母 710 其余 700
            for d in digits:
                d[0], d[1], d[2] = "7", "0", "0"
            if digits:
                digits[-1][1] = "1"
        else:
            for d in digits:
                d[0], d[1], d[2] = "7", "0", "0"

    def _apply_num_rules(self, lower: str, digits: List[List[str]]) -> None:
        n = len(digits)
        for d in digits:
            d[0], d[1], d[2] = "9", "0", "0"
        # 序数词启发式
        if re.search(r"(st|nd|rd|th)$", lower) and n >= 2:
            self._apply_emphasis(digits, [n - 2, n - 1], second="2")

    # -----------------------------
    # 后三位（语法辅助）
    # -----------------------------

    def _detect_clause_nonfinite_marks(self, tagged: Sequence[Tuple[str, str]]) -> List[str]:
        """第四位：从句/非谓语边界。
        0=无标记, 1=从句起点, 2=从句终点, 3=不定式起点, 4=VBG非谓语, 5=VBN非谓语
        """
        marks = ["0"] * len(tagged)

        for i, (tok, pos) in enumerate(tagged):
            lower = tok.lower()
            if lower in self.subordinators:
                marks[i] = "1"
            if pos == "TO" and i + 1 < len(tagged) and tagged[i + 1][1] == "VB":
                marks[i] = "3"
            if pos == "VBG":
                marks[i] = "4"
            if pos == "VBN":
                marks[i] = "5"

        # 简单从句终点：逗号/句末标点前一个词
        for i, (tok, _) in enumerate(tagged):
            if tok in {",", ";", ".", "!", "?"} and i - 1 >= 0 and marks[i - 1] == "0":
                marks[i - 1] = "2"
        return marks

    def _detect_sentence_components(self, tagged: Sequence[Tuple[str, str]]) -> List[str]:
        """第五位：句子成分。
        0=未标注, 1=主语候选, 2=谓语, 3=宾语, 4=表语, 5=状语
        """
        marks = ["0"] * len(tagged)

        first_finite = -1
        for i, (_, pos) in enumerate(tagged):
            if pos in {"VB", "VBP", "VBZ", "VBD", "MD"}:
                first_finite = i
                break

        if first_finite != -1:
            # 谓语
            marks[first_finite] = "2"

            # 主语候选：谓语前最近名词/代词短窗口
            for i in range(max(0, first_finite - 4), first_finite):
                if tagged[i][1] in {"PRP", "NN", "NNS", "NNP", "NNPS", "WP"}:
                    marks[i] = "1"

            # 宾语候选：谓语后名词/代词
            for i in range(first_finite + 1, min(len(tagged), first_finite + 6)):
                if tagged[i][1] in {"PRP", "NN", "NNS", "NNP", "NNPS"}:
                    marks[i] = "3"

            # 系动词后形容词 -> 表语
            if tagged[first_finite][0].lower() in self.copulas:
                for i in range(first_finite + 1, min(len(tagged), first_finite + 4)):
                    if tagged[i][1].startswith("JJ"):
                        marks[i] = "4"

        # 状语：副词/介词
        for i, (_, pos) in enumerate(tagged):
            if pos in {"RB", "RBR", "RBS", "WRB", "IN", "TO"} and marks[i] == "0":
                marks[i] = "5"

        return marks

    def _detect_grammar_aux_marks(self, tagged: Sequence[Tuple[str, str]]) -> List[str]:
        """第六位：语法辅助。
        0=无, 1=被动语态线索(VBN且前面be), 2=情态动词, 3=虚拟语气线索
        """
        marks = ["0"] * len(tagged)

        for i, (tok, pos) in enumerate(tagged):
            lower = tok.lower()
            if pos == "MD":
                marks[i] = "2"

            if pos == "VBN" and i - 1 >= 0 and tagged[i - 1][0].lower() in self.copulas:
                marks[i] = "1"

            # 虚拟语气线索：if + were / suggest|insist + that + VB
            if lower == "were" and i - 1 >= 0 and tagged[i - 1][0].lower() == "if":
                marks[i] = "3"
            if lower in {"suggest", "insist", "recommend", "demand"}:
                for j in range(i + 1, min(len(tagged), i + 5)):
                    if tagged[j][0].lower() == "that":
                        marks[j] = "3"

        return marks

    # -----------------------------
    # 语境判定
    # -----------------------------

    def _is_gerund_context(self, tagged: Sequence[Tuple[str, str]], index: int) -> bool:
        """VBG 是否更像动名词：前面是介词/限定词/所有格时倾向动名词。"""
        if index == 0:
            return False
        prev_word, prev_pos = tagged[index - 1]
        if prev_pos in {"IN", "DT", "PRP$", "POS"}:
            return True
        if prev_word.lower() in {"by", "for", "of", "about", "before", "after"}:
            return True
        return False


def demo() -> None:
    sample_text = """Lesson 1 A private conversation
Last week I went to the theatre.
I did not enjoy it.
' What a day ! ' I thought.
"""
    encoder = VisualGrammarEncoder()

    if nltk is None:
        print("当前环境未安装 NLTK：仅演示解码接口与 docx 写入结构。")
        compact_demo = "d000100o000100"
        print("decode_compact_token:", encoder.decode_compact_token(compact_demo))
        saved = encoder._write_simple_docx(["NLTK unavailable demo"], "encoded_output.docx")
        print(f"Word 文件已保存: {saved}")
        return

    encoded = encoder.encode_text(sample_text)
    for item in encoded[:30]:
        if item.letters and item.letters[0].code == "":
            print(f"{item.token}/{item.pos}: {item.token}")
        else:
            print(f"{item.token}/{item.pos}: {item.compact}")

    print("\n" + "\n".join(encoder.format_noun_phrase_report(sample_text)))

    saved = encoder.save_encoded_text_to_word(sample_text, "encoded_output.docx")
    print(f"\nWord 文件已保存: {saved}")


if __name__ == "__main__":
    demo()
